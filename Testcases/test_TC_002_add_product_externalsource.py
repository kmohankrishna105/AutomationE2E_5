#xlrd,xlwt----Advanced : openpyxl
import openpyxl
from Base import InitiateDriver
from Pages import LoginPage
from Lib import utils

def Data_generate():
    """
    1.Load a work book
    2.Read the excel sheet
    3.How many records are there in excel
    4.iterating til the last row

    :return the list if user details:
    """
    wb=openpyxl.load_workbook("C:\\Users\\user\\PycharmProjects\\AutomationE2E\\Data\\user_data.xlsx")
    print(wb.style_names)
    sheet=wb.get_sheet_by_name('Sheet1')
    rows=sheet.max_row
    li=[]
    lis=[]
    for i in range(1,rows+1):
        lis=[]
        user=sheet.cell(i,1)
        password=sheet.cell(i,2)
        #print(user,password)
        lis.insert(0,user.value)
        lis.insert(1,password.value)
        li.insert(i-1,lis)
    print(li)
    return li


import pytest
@pytest.mark.parametrize('data',Data_generate())
def test_login(data):
    driver=InitiateDriver.startBrowser()
    login=LoginPage.LoginClass(driver)
    login.signinClick()
    login.enter_email(data[0])
    login.enter_password(data[1])
    login.click_submit()
    from Pages import Product
    product=Product.ProductClass(driver)
    product.Click_women_link()
    utils.Take_Screenshot(driver,data[0]+"_women")
    product.select_product()
    utils.Take_Screenshot(driver, data[0]+"women_sub")
    product.select_subproduct()
    product.select_size("L")
    utils.Take_Screenshot(driver,"size_selection")


